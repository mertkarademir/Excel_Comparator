# Excel comparator using python by mert karademir
# 29 October 2023
# This script takes 3 excel files and compare every cells with eachother, the different cell will be displayed 

import openpyxl
import os
import pygame

# playing an evil opening theme from Phineas and Ferb. You can customize by adding any .mp3 
pygame.mixer.init()
pygame.mixer.music.load('doof.mp3')
pygame.mixer.music.play()

# Acquiring the excel files from the folder
def get_excel_files_in_directory(directory):
    excel_files = [file for file in os.listdir(directory) if file.endswith('.xlsx')]
    return excel_files

# Comparing the cells of excel files.
def compare_worksheets(ws1, ws2, ws3):
    differences = []
    for row1, row2, row3 in zip(ws1.iter_rows(), ws2.iter_rows(), ws3.iter_rows()):
        for cell1, cell2, cell3 in zip(row1, row2, row3):
            values = [cell1.value, cell2.value, cell3.value]
            unique_values = list(set(values))
            if len(unique_values) != 1:
                different_value = [v for v in unique_values if values.count(v) == 1][0]
                different_file = excel_files[values.index(different_value)]
                differences.append((cell1.coordinate, different_file, different_value))
    return differences

directory = "./"
excel_files = get_excel_files_in_directory(directory)

if len(excel_files) != 3:
    print("You need to have exactly 3 Excel files in the directory.")
else:
    wb1 = openpyxl.load_workbook(excel_files[0])
    wb2 = openpyxl.load_workbook(excel_files[1])
    wb3 = openpyxl.load_workbook(excel_files[2])

    differences = compare_worksheets(wb1.active, wb2.active, wb3.active)

    def print_differences(differences):
        for diff in differences:
            cell_coord, file_name, cell_value = diff
            print(f"Different Value in Cell {cell_coord} in {file_name}: {cell_value}")

    if differences:
        print("Wrong Answers in File1, File2, and File3:")
        print_differences(differences)
    else:
        print("All Answers are Different (Wrong) in File1, File2, and File3.")

exit = input("Enter any value to close the program")
while pygame.mixer.music.get_busy():
    pygame.time.Clock().tick(10)

# To make this script an executable file,
# python -m PyInstaller ./excel_comparator.py --onefile
